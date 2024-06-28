from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
import pyautogui
import pyinput
import io
import tempfile
import time
from PIL import Image
import os

navegador = webdriver.Chrome()
navegador.get("https://bedinsat.com.br/panel/_login.php")
navegador.find_element("xpath", '/html/body/div[1]/div/div[2]/div/div[2]/form/div/ul/li[3]/input').send_keys("**##@@@##**")
navegador.find_element("xpath", '/html/body/div[1]/div/div[2]/div/div[2]/form/div/ul/li[5]/input').send_keys("**##@@@##**")
botao_entrar = navegador.find_element("xpath", '/html/body/div[1]/div/div[2]/div/div[2]/form/div/ul/li[7]/input')
botao_entrar.click()
time.sleep(2)

botao_Produtos = navegador.find_element("xpath", '/html/body/div[2]/div/div/div/ul/li[4]/a')
botao_Produtos.click()
botao_Produtos2 = navegador.find_element("xpath", '/html/body/div[2]/div/div/div/ul/li[4]/ul/li[1]/a')
botao_Produtos2.click()
time.sleep(2)
arquivo_titulo_br = "C:/Users/ti3/OneDrive/Área de Trabalho/autoPython/Titulos/tituloBR.txt" 
arquivo_titulo_en = "C:/Users/ti3/OneDrive/Área de Trabalho/autoPython/Titulos/tituloEN.txt" 
arquivo_titulo_es = "C:/Users/ti3/OneDrive/Área de Trabalho/autoPython/Titulos/tituloES.txt"

arquivo_codigo =  "C:/Users/ti3/OneDrive\Área de Trabalho/Atualizado/Scraping and Writing/autoPython/Codes py/codigos.txt"

arquivo_descricao_br = "C:/Users/ti3/OneDrive/Área de Trabalho/Atualizado/Scraping and Writing/descricoesBR.xlsx"
arquivo_descricao_en = "C:/Users/ti3/OneDrive/Área de Trabalho/Atualizado/Scraping and Writing/descricoesEN.xlsx"
arquivo_descricao_es = "C:/Users/ti3/OneDrive/Área de Trabalho/Atualizado/Scraping and Writing/descricoesES.xlsx"

wb_descricao_br = load_workbook(arquivo_descricao_br)
wb_descricao_en = load_workbook(arquivo_descricao_en)
wb_descricao_es = load_workbook(arquivo_descricao_es)

ws_descricao_br = wb_descricao_br.active
ws_descricao_en = wb_descricao_en.active
ws_descricao_es = wb_descricao_es.active
linha_atual = 1
ordem = 120

while True:
    botao_Adicionar = navegador.find_element("xpath", '/html/body/div[3]/a[1]')
    botao_Adicionar.click()
    time.sleep(1)

    with open(arquivo_titulo_br, 'r', encoding='utf-8') as arquivoTitleBR, \
        open(arquivo_titulo_en, 'r', encoding='utf-8') as arquivoTitleEN, \
        open(arquivo_titulo_es, 'r', encoding='utf-8') as arquivoTitleES, \
        open(arquivo_descricao_br, 'r', encoding='utf-8') as arquivoDescBR, \
        open(arquivo_descricao_en, 'r', encoding='utf-8') as arquivoDescEN, \
        open(arquivo_descricao_es, 'r', encoding='utf-8') as arquivoDescES, \
        open(arquivo_codigo, 'r', encoding='utf-8') as arquivoCodigos:

        for linha in arquivoTitleBR:
            tituloBR = linha.strip()
            tituloEN = arquivoTitleEN.readline().strip()
            tituloES = arquivoTitleES.readline().strip()
            
            codigos = arquivoCodigos.readline().strip()

            descricaoBR = ws_descricao_br.cell(row=linha_atual, column=1).value
            descricaoEN = ws_descricao_en.cell(row=linha_atual, column=1).value
            descricaoES = ws_descricao_es.cell(row=linha_atual, column=1).value
            
            if not tituloBR and not tituloEN and not tituloES and not codigos:
                break
            
            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[1]/td[2]/input[1]').send_keys(tituloBR)
            #navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[1]/td[2]/input[2]').send_keys(tituloEN)
            #navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[1]/td[2]/input[3]').send_keys(tituloES)
            
            time.sleep(2)

            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[2]/td[2]/input[1]').send_keys(tituloBR)
            #navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[2]/td[2]/input[2]').send_keys(tituloEN)
            #navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[2]/td[2]/input[3]').send_keys(tituloES)
            
            time.sleep(2)
             
            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[3]/td[2]/input').send_keys(codigos)

            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[4]/td[2]/div[3]/label').click()
            
            time.sleep(2)
            
            iframe = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'iframe[title="Editor de Texto,pro_descricao_br"]')))
            navegador.switch_to.frame(iframe)
            html_tag = navegador.find_element(By.TAG_NAME,'html')
            body_tag = html_tag.find_element(By.TAG_NAME, 'body')
            p_tag = body_tag.find_element(By.TAG_NAME, 'p')
            p_tag.send_keys(descricaoBR)
            
            navegador.switch_to.default_content()
            
            iframe = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'iframe[title="Editor de Texto,pro_descricao_en"]')))
            navegador.switch_to.frame(iframe)
            html_tag = navegador.find_element(By.TAG_NAME,'html')
            body_tag = html_tag.find_element(By.TAG_NAME, 'body')
            p_tag = body_tag.find_element(By.TAG_NAME, 'p')
            #p_tag.send_keys(descricaoEN)
            
            navegador.switch_to.default_content()
            
            iframe = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'iframe[title="Editor de Texto,pro_descricao_es"]')))
            navegador.switch_to.frame(iframe)
            html_tag = navegador.find_element(By.TAG_NAME,'html')
            body_tag = html_tag.find_element(By.TAG_NAME, 'body')
            p_tag = body_tag.find_element(By.TAG_NAME, 'p')
            #p_tag.send_keys(descricaoES)
            
            navegador.switch_to.default_content()
            
            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[6]/td[2]/input').send_keys('academia')
            
            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[7]/td[2]/input').send_keys(ordem)

            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/table/tbody/tr[8]/td[2]/label/input').click()
            
            time.sleep(1000000)            

            navegador.find_element("xpath", '/html/body/div[4]/div/div[2]/div[1]/form/div/button[1]').click()
            
            time.sleep(2)
            
            navegador.find_element("xpath", '/html/body/div[4]/table/tbody/tr[1]/td[2]/a/img').click()
            time.sleep(1)
            navegador.find_element("xpath", '/html/body/div[4]/div[3]/div[4]/form/div[1]').click()
            time.sleep(1.5)
            pyautogui.press('f4')
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.write('C:/Users/ti3/OneDrive/imagens.jpeg')
            pyautogui.press('enter')
            time.sleep(0.5)
            pyautogui.hotkey('f3')
            pyautogui.write(tituloEN)
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(0.5)
            pyautogui.press('down')
            time.sleep(0.5)
            pyautogui.press('space')
            time.sleep(0.5)
            pyautogui.press('enter')
            navegador.find_element("xpath", '/html/body/div[4]/div[3]/div[5]/div/table/tbody/tr/td/a').click()
            
            arquivoTitleBR.readline()
            arquivoTitleEN.readline()
            arquivoTitleES.readline()
            arquivoCodigos.readline()
            linha_atual += 1
            ordem += 1
