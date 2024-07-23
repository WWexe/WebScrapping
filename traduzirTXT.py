from openpyxl import Workbook

def texto_para_excel(arquivo_texto, arquivo_excel):
    wb = Workbook()
    ws = wb.active
    
    with open(arquivo_texto, 'r', encoding='latin-1') as file:
        linhas = file.readlines()
    
    for i, linha in enumerate(linhas, start=1):
        # Dividir o texto em blocos separados por quebra de linha
        blocos = linha.split('\n')
        
        # Escrever cada bloco em células separadas na mesma linha
        for j, bloco in enumerate(blocos, start=1):
            ws.cell(row=i, column=j, value=bloco)
    
    wb.save(arquivo_excel)
    print(f"Arquivo Excel '{arquivo_excel}' gerado com sucesso!")

arquivo_texto = "C:/Users/ti3/OneDrive/Área de Trabalho/Atualizado/Scraping and Writing/descricoesES.txt"

arquivo_excel = "C:/Users/ti3/OneDrive/Área de Trabalho/Atualizado/Scraping and Writing/descricoesES.xlsx"

texto_para_excel(arquivo_texto, arquivo_excel)

from openpyxl import load_workbook

def apagar_celulas_vazias_coluna1(arquivo_excel):
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                ws.delete_rows(cell.row, amount=1)
    
    wb.save(arquivo_excel)
    print("Células vazias da coluna 1 removidas com sucesso!")

arquivo_excel = "C:/Users/ti3/OneDrive/Área de Trabalho/Atualizado/Scraping and Writing/descricoesES.xlsx"

apagar_celulas_vazias_coluna1(arquivo_excel)


